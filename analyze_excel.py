import openpyxl
import json
import os

base = r'C:\Antigravity 1\Productive'
files_to_check = [
    'Odatlar 2026.xlsx',
    'Produktivlik 2026.xlsx',
    'Vazifalar 2026.xlsx',
]

# Check for Cyrillic filename
for f in os.listdir(base):
    if f.endswith('.xlsx') and f not in files_to_check:
        files_to_check.append(f)

result = {}

for f in files_to_check:
    fp = os.path.join(base, f)
    if not os.path.exists(fp):
        result[f] = {'error': 'File not found'}
        continue
    try:
        wb = openpyxl.load_workbook(fp, data_only=True)
        file_info = {'sheets': list(wb.sheetnames), 'details': {}}
        for sn in wb.sheetnames:
            ws = wb[sn]
            rows_data = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), 
                                     max_col=min(ws.max_column, 30), values_only=False):
                cells = []
                for cell in row:
                    v = cell.value
                    if v is not None:
                        cells.append({
                            'col': cell.column,
                            'row': cell.row,
                            'val': str(v)[:80],
                            'merged': isinstance(cell, openpyxl.cell.cell.MergedCell)
                        })
                if cells:
                    rows_data.append(cells)
            file_info['details'][sn] = {
                'max_row': ws.max_row,
                'max_col': ws.max_column,
                'merged_cells': [str(m) for m in ws.merged_cells.ranges],
                'data': rows_data
            }
        result[f] = file_info
        wb.close()
    except Exception as e:
        result[f] = {'error': str(e)}

output_path = os.path.join(base, 'site', 'excel_analysis.json')
with open(output_path, 'w', encoding='utf-8') as fh:
    json.dump(result, fh, ensure_ascii=False, indent=2)

print("Analysis complete. Output saved to:", output_path)
print("\n=== SUMMARY ===")
for fname, info in result.items():
    if 'error' in info:
        print(f"\n{fname}: ERROR - {info['error']}")
    else:
        print(f"\n{fname}:")
        print(f"  Sheets: {info['sheets']}")
        for sn, details in info['details'].items():
            print(f"  [{sn}] rows={details['max_row']}, cols={details['max_col']}, merged={len(details['merged_cells'])}")
            if details['data']:
                first_row = details['data'][0]
                headers = [c['val'] for c in first_row]
                print(f"    Headers/Row1: {headers[:15]}")
                if len(details['data']) > 1:
                    second_row = details['data'][1]
                    vals = [c['val'] for c in second_row]
                    print(f"    Row2: {vals[:15]}")
